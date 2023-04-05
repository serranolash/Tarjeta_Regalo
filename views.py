from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout
from .forms import  ImprimirEtiquetasForm, TipoForm, LocalForm,  VoucherForm,  VenderForm, VoucherUpdateForm, AuthenticationForm
from .models import Voucher, Etiqueta, Tipo, generar_tipo, Local
from django.views.generic import ListView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.views.generic import View
import barcode
from barcode import Code128
from barcode.writer import ImageWriter
from io import BytesIO
from django.core.files import File
import base64
import zebra
from .utils import generar_vouchers
from barcode import generate
from barcode.writer import ImageWriter
from io import BytesIO
from datetime import datetime
from django.utils import timezone
from django.urls import reverse
from django.http import Http404, HttpResponse
from django.views.generic import CreateView
from .models import CustomUser
from .forms import CustomUserCreationForm
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os
import io
from django.views.decorators.csrf import csrf_protect
import pytz
from django.utils import timezone
from django.conf import settings
from datetime import datetime
from django.utils import timezone
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.http import HttpResponse
from gifvoucher_app.models import Voucher
from django.shortcuts import redirect
from django.http import HttpResponse
import openpyxl
import io
import os
import xlsxwriter
from django.http import HttpResponse
from django.shortcuts import redirect
from django.views import View
from django.contrib import messages
from .forms import ImprimirEtiquetasForm
from .models import Etiqueta, Tipo
from .utils import generar_vouchers
import openpyxl
from django.http import HttpResponse
# Obtener el objeto datetime desde la base de datos
my_datetime = timezone.now()
# Convertir a la zona horaria configurada en Django
my_datetime = timezone.localtime(my_datetime, pytz.timezone(settings.TIME_ZONE))



class CustomUserCreateView(CreateView):
    model = CustomUser
    form_class = CustomUserCreationForm
    success_url = reverse_lazy('home')
    template_name = 'registration/register.html'

@login_required
def home(request):
    if not isinstance(request.user, CustomUser):
        # Si el objeto User no es una instancia de CustomUser, redirigir a una página de error o mostrar un mensaje de error
        return HttpResponse("Error: el usuario no es una instancia de CustomUser")

    if request.user.user_type == 3:  # Administrador
        # Mostrar todas las vistas del menú
        return render(request, 'menu_administrador.html')
    elif request.user.user_type == 2:  # Usuario avanzado
        # Mostrar algunas vistas del menú
        return render(request, 'menu_usuario_avanzado.html')
    else:  # Usuario común
        # Mostrar otras vistas del menú
        return render(request, 'menu_usuario_comun.html')



# Create your views here.
def bienvenida(request):
    return render(request, 'bienvenida.html')

from django.contrib.auth import authenticate, login

def registro(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Autenticar al usuario recién creado y redirigirlo al inicio
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registro.html', {'form': form})

from django.contrib.auth import authenticate, login as auth_login


def logear(request):
    if request.method == 'POST':
        print("Formulario recibido") # mensaje de depuración
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            print("Email: ", email) # mensaje de depuración
            print("Password: ", password) # mensaje de depuración
            user = authenticate(request, email=email, password=password)
            print("User: ", user) # mensaje de depuración
            if user is not None:
              login(request, user)
              if user.user_type == 1:
                print("Redirigiendo a página de usuario común")
                return redirect('menu_usuario_comun')
              elif user.user_type == 2:
                print("Redirigiendo a página de usuario avanzado")
                return redirect('menu_usuario_avanzado')
              elif user.user_type == 3:
                print("Redirigiendo a página de administrador")
                return redirect('menu_administrador')
              elif user.user_type == 4:
                print("Redirigiendo a página de usuario auditor")
                return redirect('menu_usuario_auditor')

                
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})


@login_required
def menu_administrador(request):
    return render(request, 'menu_administrador.html')

@login_required
def menu_usuario_comun(request):
    return render(request, 'menu_usuario_comun.html')

@login_required
def menu_usuario_avanzado(request):
    return render(request, 'menu_usuario_avanzado.html')

@login_required
def menu_usuario_auditor(request):
    return render(request, 'menu_usuario_auditor.html')

def logout_view(request):
    
    return redirect('login')

from .models import Voucher

def crear_voucher(request):
    if request.method == 'POST':
        form = TipoForm(request.POST)
        if form.is_valid():
            tipo = form.save(commit=False)
            tipo.tipo = generar_tipo(form.cleaned_data['monto'])
            tipo.save()
            return redirect('voucher_creado', pk=tipo.pk)
    else:
        form = TipoForm()
    return render(request, 'crear_voucher.html', {'form': form})

def voucher_creado(request, pk):
    tipo = get_object_or_404(Tipo, pk=pk)
    return render(request, 'voucher_creado.html', {'tipo': tipo})


@method_decorator(login_required, name='dispatch')
class VoucherListView(ListView):
    model = Voucher
    template_name = 'voucher_list.html'
    context_object_name = 'vouchers'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ImprimirEtiquetasForm()
        return context

import openpyxl
from django.http import HttpResponse


class GenerarEtiquetasView(View):
    def get(self, request):
        form = ImprimirEtiquetasForm()
        tipo = Tipo.objects.all()
        return render(request, 'imprimir_etiquetas.html', {'form': form, 'tipos': tipo})

    def post(self, request, *args, **kwargs):
        form = ImprimirEtiquetasForm(request.POST)
        if form.is_valid():
            tipo = form.cleaned_data['tipo']
            cantidad = form.cleaned_data['cantidad']
            usuario= request.user.username
            vouchers = generar_vouchers(tipo, cantidad, usuario)
            etiquetas_generadas = []
            for voucher in vouchers:
                etiqueta = Etiqueta.objects.create(serie=voucher.serie, creado_usr=usuario)
                etiquetas_generadas.append(etiqueta)              
                print(etiqueta)
            messages.success(request, 'Etiquetas creadas satisfactoriamente.')
            return render(request, 'imprimir_etiquetas.html', {'form': form, 'etiquetas_generadas': etiquetas_generadas})
        else:
            messages.error(request, "Por favor corrija los errores del formulario")
        tipo = Tipo.objects.all()
        return render(request, 'imprimir_etiquetas.html', {'form': form, 'tipos': tipo})


def lista_etiquetas(request):
    etiquetas_con_imagenes = []

    # Buscar etiquetas que coincidan con el número de serie especificado
    serie = request.GET.get('serie')
    if serie:
        etiquetas = Etiqueta.objects.filter(serie__icontains=serie)        
    else:
        etiquetas = Etiqueta.objects.all()

    for etiqueta in etiquetas:
        buffer = BytesIO()
        generate('Code128', etiqueta.serie, writer=ImageWriter(), output=buffer)
        barcode_blob = buffer.getvalue()

        imagen_base64 = base64.b64encode(barcode_blob).decode('ascii')
        etiquetas_con_imagenes.append({
            'id': etiqueta.id,
            'serie': etiqueta.serie,
            'imagen': 'data:image/png;base64,' + imagen_base64,
            'creado_fecha': etiqueta.creado_fecha.strftime('%Y-%m-%d %H:%M:%S')
        })

   
    return render(request, 'lista_etiquetas.html', {'etiquetas': etiquetas_con_imagenes})

def generacion_impresiones(request):
    etiquetas_con_imagenes = []
    if request.method == 'POST':
        inicio = request.POST.get('inicio')
        fin = request.POST.get('fin')
        print(f"inicio: {inicio}")
        print(f"fin: {fin}")
        usuario = request.user.username
        
        # Obtener los vouchers a imprimir
        vouchers = Voucher.objects.filter(serie__gte=inicio, serie__lte=fin)
        print(f"Vouchers encontrados: {vouchers.count()}")

        # Actualizar los campos de la base de datos
        num_vouchers_actualizados = 0
        for voucher in vouchers:
            if not voucher.impreso:
                voucher.impresiones += 1
                voucher.impreso_fecha = timezone.now()
                voucher.impreso_usr = usuario
                voucher.impreso = True
            else:
                voucher.impresiones += 1
                voucher.impreso_fecha = timezone.now()
                voucher.impreso_usr = usuario
            try:
                voucher.save()
                num_vouchers_actualizados += 1
            except Exception as e:
                print(f"Error al guardar voucher: {e}")
                messages.error(request, f"Error al guardar voucher: {e}")

        # Verificar si se actualizaron los vouchers y mostrar un mensaje
        if num_vouchers_actualizados > 0:
            messages.success(request, f'Se actualizaron {num_vouchers_actualizados} vouchers correctamente en la base de Datos y fueron mandados a imprimir.')
        else:
            messages.warning(request, 'No se encontraron vouchers para actualizar.')
        
        # Crear diccionario con datos de vouchers actualizados
        datos_actualizados = {}
        for voucher in vouchers:
            if voucher.impreso:
                datos_actualizados[voucher.serie] = {
                    'impresiones': voucher.impresiones,
                    'impreso_fecha': voucher.impreso_fecha,
                    'impreso_usr': voucher.impreso_usr,
                }

        # Pasar diccionario al contexto de la plantilla
        contexto = {
            'datos_actualizados': datos_actualizados,
        }

        return render(request, 'generacion_impresiones.html', contexto)
    
    return render(request, 'generacion_impresiones.html')

from django.shortcuts import render
from django.http import HttpResponseRedirect
from .models import Voucher, Local

def asignar_vouchers(request):
    locales = Local.objects.all()
    context = {'locales': locales}    
    voucher_error = None  # variable para almacenar un mensaje de error si el voucher ya ha sido asignado
    voucher_local_error = None  # variable para almacenar un mensaje de error si el voucher ya ha sido recibido en otro local
    if request.method == 'POST':
        desde = request.POST['desde']
        hasta = request.POST['hasta']
        local_id = request.POST['local']
        local = Local.objects.get(id=local_id)
        
        # Verificar si el voucher ingresado existe en la base de datos
        voucher_serie = request.POST.get('voucher')
        if voucher_serie:
            try:
                voucher = Voucher.objects.get(serie=voucher_serie)
            except Voucher.DoesNotExist:
                voucher_error = 'El voucher ingresado no es válido. Ingrese un voucher válido.'
                context = {'voucher_error': voucher_error}
                return render(request, 'asignar_vouchers.html', context)
            
            # Verificar si el voucher ha sido recibido
            if voucher.recibido:
                # Verificar si el voucher ha sido recibido en el local actual
                if voucher.local == local:
                    voucher_error = 'El voucher ya ha sido asignado. No se puede asignar nuevamente.'
                else:
                    voucher_local_error = 'El voucher ya ha sido recibido en otro local. No se puede asignar.'
                context = {'voucher_error': voucher_error, 'voucher_local_error': voucher_local_error}
                return render(request, 'asignar_vouchers.html', context)

        # Actualizar los vouchers en la base de datos
        vouchers = Voucher.objects.filter(serie__gte=desde, serie__lte=hasta, recibido=False)
        vouchers.update(enviado=True, enviado_fecha=timezone.now(), enviado_usr=request.user.username, local=local)

        messages.success(request, 'Vouchers asignados correctamente.')
        return redirect(reverse('asignar_vouchers'))

    # Si la solicitud es GET, muestra la página de asignación de vouchers
    locales = Local.objects.all()
    context = {'locales': locales, 'voucher_error': voucher_error, 'voucher_local_error': voucher_local_error}
    return render(request, 'asignar_vouchers.html', context)

from django.contrib import messages

def elegir_local(request):
    local_form = LocalForm(request.POST or None)
    voucher_form = VoucherForm(request.POST or None)

    if local_form.is_valid():
        print("Local form is valid")
        local_id = local_form.cleaned_data['local_id']
        vouchers = Voucher.objects.filter(local=local_id).exclude(recibido=True)
        voucher_form.fields['vouchers'].queryset = vouchers

        if voucher_form.is_valid():
            print("Voucher form is valid")
            selected_vouchers = voucher_form.cleaned_data.get('vouchers')
            for voucher in selected_vouchers:
                if voucher.recibido:
                    messages.warning(request, f"El voucher {voucher} ya fue recibido.")
                else:
                    voucher.recibido = True
                    voucher.recibido_fecha = timezone.now()
                    voucher.recibido_usr = request.user.username
                    voucher.save()
                    print(voucher.recibido)
                    print(voucher.recibido_fecha)
                    print(voucher.recibido_usr)

            # Mostrar un mensaje de agradecimiento dentro de la misma plantilla
            context = {
                'local_form': local_form,
                'voucher_form': voucher_form,
                'formulario_valido': True,
            }
            return render(request, 'elegir_local.html', context)

    context = {
        'local_form': local_form,
        'voucher_form': voucher_form,
    }
    return render(request, 'elegir_local.html', context)   

def vender_voucher(request):
    serie = request.POST.get('serie', '')    
    vouchers = Voucher.objects.filter(estado='DISPONIBLE', serie=serie)

    context = {'vouchers': vouchers, 'serie': serie}

    if request.method == 'POST' and 'voucher_serie' in request.POST:
        voucher_serie = request.POST.get('voucher_serie')
        voucher = Voucher.objects.get(serie=voucher_serie, estado='DISPONIBLE')
        voucher.vendido = True
        voucher.vendido_fecha = timezone.now()
        voucher.vendido_usr = request.user.username
        voucher.save()
        context['venta_exitosa'] = f'Se ha vendido el voucher {voucher_serie} exitosamente.'

    return render(request, 'vender_voucher.html', context)

def usar_voucher(request):
    vouchers = Voucher.objects.filter(estado='DISPONIBLE').order_by('serie')
    if request.method == 'POST':
        try:
            voucher_series = request.POST.getlist('voucher_seleccionado')
            vouchers = Voucher.objects.filter(serie__in=voucher_series, vendido=True)
            if vouchers.count() == len(voucher_series):
                vouchers.update(
                    usado=True,
                    usado_fecha=timezone.now(),
                    usado_usr=request.user.username,
                    estado='NO DISPONIBLE'
                )
                messages.success(request, 'Voucher(s) utilizado(s) exitosamente.')
            else:
                messages.error(request, 'Uno o más vouchers seleccionados no han sido vendidos.')
            return redirect(reverse('usar_voucher'))
        except Voucher.DoesNotExist:
            messages.error(request, 'No se encontró el voucher seleccionado.')
    return render(request, 'usar_voucher.html', {'vouchers': vouchers})

import pytz

import pytz
from django.shortcuts import render
from .models import Voucher


def update_voucher(request):
    mensaje = None
    voucher = None
    campos_actualizados_str = ""

    if request.method == 'POST':
        # obtener el número de serie del voucher enviado por el usuario
        serie_voucher = request.POST.get('serie_voucher')
        print("Número de serie del voucher enviado por el usuario:", serie_voucher)

        # buscar los vouchers que tengan el campo "usado" marcado como verdadero
        vouchers_usados = Voucher.objects.filter(usado=True)
        print("Vouchers usados encontrados:", vouchers_usados)

        # crear una lista de los números de serie de los vouchers usados
        series_vouchers_usados = [voucher.serie for voucher in vouchers_usados]
        print("Lista de números de serie de vouchers usados:", series_vouchers_usados)

        # verificar si el número de serie del voucher enviado por el usuario está en la lista de números de serie de vouchers usados
        if serie_voucher in series_vouchers_usados:
            # Obtener la zona horaria actual
            tz = pytz.timezone('America/Lima')

            # Crear un objeto datetime con la zona horaria
            now = datetime.now(tz)
            
            # si el número de serie está en la lista y no ha sido auditado, actualizar los campos de "voucher"
            voucher = Voucher.objects.get(serie=serie_voucher)
            if voucher.audit == False:
                campos_actualizados = []
                if voucher.audit != True:
                    voucher.audit = True
                    voucher.audit_fecha = now  # importar datetime al principio del archivo
                    voucher.audit_usr = request.user.username
                    campos_actualizados.append("audit")
                if voucher.destr != True:
                    voucher.destr = True
                    voucher.destr_fecha = now
                    voucher.destr_usr = request.user.username
                    campos_actualizados.append("destr")
                voucher.save()
                mensaje_debug = f"Voucher guardado correctamente: {voucher.serie}"
                print(mensaje_debug)

                mensaje = "¡El voucher se ha actualizado correctamente!"
                print("Voucher actualizado correctamente:", voucher)

                # crear una lista de los campos que se actualizaron para mostrarlos en la plantilla
                campos_actualizados_str = ",".join(campos_actualizados)
            else:
                mensaje = "El número de serie del voucher ya ha sido auditado."
                print("El número de serie del voucher ya ha sido auditado.")
        else:
            mensaje = "El número de serie del voucher no está marcado como usado."
            print("El número de serie del voucher no está marcado como usado.")

    return render(request, 'update_voucher.html', {'mensaje': mensaje, 'voucher': voucher, 'campos_actualizados': campos_actualizados_str})

from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Voucher

def voucher_listado(request):
    vouchers = Voucher.objects.all()
    context = {'vouchers': vouchers}
    return render(request, 'voucher_listado.html', context)

def export_vouchers(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="vouchers.xlsx"'

    wb = Workbook()
    ws = wb.active

    # Headers
    headers = [
        'Serie', 'Lote', 'Orden', 'Tipo', 'Fecha de creación', 'Creado por', 'Impreso', 'Fecha de impresión',
        'Impreso por', 'Impresiones', 'Enviado', 'Fecha de envío', 'Enviado por', 'Vendido', 'Fecha de venta',
        'Vendido por', 'Usado', 'Fecha de uso', 'Usado por', 'Auditado', 'Fecha de auditoría', 'Auditado por',
        'Destruido', 'Fecha de destrucción', 'Destruido por', 'Monto', 'Moneda', 'Estado', 'Local', 'Recibido',
        'Fecha de recepción', 'Recibido por'
    ]
    for col_num, header_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.cell(row=1, column=col_num, value=header_title)
        ws.column_dimensions[col_letter].width = 20

    # Rows
    vouchers = Voucher.objects.all()
    for row_num, voucher in enumerate(vouchers, 1):
        row = [
    voucher.serie, voucher.lote, voucher.orden, voucher.tipo,
    voucher.creado_fecha.replace(tzinfo=None) if voucher.creado_fecha else None,
    voucher.creado_usr, voucher.impreso,
    voucher.impreso_fecha.replace(tzinfo=None) if voucher.impreso_fecha else None,
    voucher.impreso_usr, voucher.impresiones, voucher.enviado,
    voucher.enviado_fecha.replace(tzinfo=None) if voucher.enviado_fecha else None,
    voucher.enviado_usr, voucher.vendido,
    voucher.vendido_fecha.replace(tzinfo=None) if voucher.vendido_fecha else None,
    voucher.vendido_usr, voucher.usado,
    voucher.usado_fecha.replace(tzinfo=None) if voucher.usado_fecha else None,
    voucher.usado_usr, voucher.audit,
    voucher.audit_fecha.replace(tzinfo=None) if voucher.audit_fecha else None,
    voucher.audit_usr, voucher.destr,
    voucher.destr_fecha.replace(tzinfo=None) if voucher.destr_fecha else None,
    voucher.recibido_fecha.replace(tzinfo=None) if voucher.recibido_fecha else None,
    voucher.destr_usr, voucher.monto, voucher.moneda, voucher.estado, str(voucher.local), voucher.recibido, voucher.recibido_usr
]

        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            ws.cell(row=row_num+1, column=col_num, value=cell_value)

    wb.save(response)
    return response

